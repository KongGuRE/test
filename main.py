import openpyxl

if __name__ == '__main__':
    # 엑셀 파일 만들기
    # wb = openpyxl.Workbook()
    # wb.create_sheet(title='NewSheet')
    # wb.save('test.xlsx')

    # 엑셀 파일 오픈
    filename = "test.xlsx"
    wb = openpyxl.load_workbook(filename)

    ## 시트 설정
    print(wb.worksheets)
    sheet = wb.worksheets[0] # active 시트 열기
    # sheet = wb['Sheet']  # 시트명 직접 적어주기

    ## 데이터 가져오기
    rowCount = 2
    for row in sheet.rows:
        print(row)
        try:
            ## 엑셀 읽어오기
            read_cell = row[0].value
            print(read_cell)

            ## cell 설정
            lat_cell1 = sheet.cell(row=rowCount, column=3, value="위도")  # C열은 3
            lng_cell2 = sheet.cell(row=2, column=1, value="경도")

            wb.save(filename=filename)

        except KeyError as ke:
            None
            # lat_cell.value = 0
            # lng_cell.value = 0

            "test1 version"
